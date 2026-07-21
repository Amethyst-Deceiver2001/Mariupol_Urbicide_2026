#!/usr/bin/env python3
"""Parse the AGO Mariupol official compensation-housing lists captured by
scripts/392 (source_types ago_mariupol_housing_distribution_xlsx /
ago_mariupol_housing_queue_xlsx). Read-only over the raw store; writes
data/parsed/ago_lost_dwellings.jsonl. No network, no DB writes.

WHAT THESE LISTS ARE. Each distribution row is one compensation RECIPIENT,
anonymised at source to a 6-hex ID, plus the address of the dwelling they
LOST ("Адрес утраченного жилья") and its district. The municipal flat they
were actually given is NOT in the file (recipient privacy). So the parseable,
spine-relevant content is the set of officially-recorded LOST/destroyed
dwellings that the compensation programme is built on -- the occupation
administration's own apartment-level admission of loss (RD4U A3.1), dated and
tied to Постановление ГКО №175 / Решение №61-1. This is the SUPPLY-SIDE
counterpart to the crowd-sourced demand-side reallocation ledger (scripts/391):
that ledger names the SEIZED flats being handed out; this list names the LOST
homes of the people receiving them. They do not join per-record (the recipient
is anonymised) but together they bound both ends of the redistribution machine.

The queue file carries no address (ID / position / decree / room-count only);
we parse it for the decree-basis split and total demand size, aggregate only.

PRIVACY (CLAUDE.md): the recipient is a living private individual, already
pseudonymised to a hex ID at source. We store the hex ID, the structured lost-
dwelling address (building/apt), district, list date and decree basis -- never
a name (there are none in the file). Shared output aggregates to building level.

    PYTHONPATH=src .venv312/bin/python scripts/394_parse_ago_distribution_lists.py
"""
from __future__ import annotations

import json
import logging
import re
import sqlite3
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

import openpyxl  # noqa: E402

from mariupol_seizures import config  # noqa: E402
from mariupol_seizures.normalize.address import address_to_building_key  # noqa: E402

log = logging.getLogger(__name__)

# "РФ, ДНР, Г. МАРИУПОЛЬ, УЛ. КУПРИНА, Д. 3, КВ. 135" -> street / house / apt.
# The street token keeps its type marker (УЛ./ПР-КТ/Б-Р), which the normalizer
# needs to classify STREET/AVENUE/BOULEVARD (see scripts/391 lesson).
ADDR_RE = re.compile(
    r"МАРИУПОЛЬ\s*,\s*(.+?)\s*,\s*Д(?:ОМ)?\.?\s*(\d+(?:[/\-]\d+)*[А-Яа-я]?)"
    r"(?:\s*,\s*КВ\.?\s*(\d+[А-Яа-я]?))?", re.I)
# collapse an accidental doubled type token ("Б-Р Б-Р ХМЕЛЬНИЦКОГО")
DUP_TYPE_RE = re.compile(r"^\s*(УЛ|ПР-?КТ|Б-?Р|ПЕР|ПРОСП|ШОССЕ)\.?\s+\1\.?\s+", re.I)

DATE_RE = re.compile(r"(\d{2}\.\d{2}\.\d{4})")


def _list_date(title: str) -> str | None:
    m = DATE_RE.search(title or "")
    if not m:
        return None
    d, mo, y = m.group(1).split(".")
    return f"{y}-{mo}-{d}"


def _rows(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    for row in it:
        yield row
    wb.close()


def parse_distribution(cur) -> list[dict]:
    out: list[dict] = []
    seen: set[tuple] = set()          # (list_date, anon_id) -> keep once per snapshot
    unparsed = 0
    docs = cur.execute(
        "SELECT sha256, title, raw_path FROM source_document "
        "WHERE source_type='ago_mariupol_housing_distribution_xlsx'",
    ).fetchall()
    for sha, title, path in docs:
        ld = _list_date(title)
        n = 0
        for row in _rows(path):
            if not row or not row[0]:
                continue
            anon_id = str(row[0]).strip()
            addr = str(row[1] or "").strip()
            district = str(row[2] or "").strip() if len(row) > 2 else ""
            if (ld, anon_id) in seen:
                continue
            seen.add((ld, anon_id))
            m = ADDR_RE.search(addr)
            if not m:
                unparsed += 1
                continue
            street = DUP_TYPE_RE.sub(lambda x: x.group(1) + " ", m.group(1)).strip(" .,")
            house, apt = m.group(2).strip(), (m.group(3) or None)
            building_id = address_to_building_key(street, house)
            if building_id is None or building_id.startswith("UNKNOWN:"):
                unparsed += 1
                continue
            out.append({
                "list_date": ld,
                "anon_id": anon_id,
                "building_id": building_id,
                "street_raw": street,
                "house_raw": house,
                "apt_raw": apt,
                "district": district,
                "source_sha256": sha,
                "source_type": "ago_mariupol_housing_distribution_xlsx",
            })
            n += 1
        log.info("distribution %s (%s): %d rows parsed", title, ld, n)
    log.info("distribution: %d parsed rows, %d unparsed/unclassifiable", len(out), unparsed)
    return out


def parse_queue(cur) -> dict:
    by_decree: Counter = Counter()
    by_rooms: Counter = Counter()
    total = 0
    doc = cur.execute(
        "SELECT title, raw_path FROM source_document "
        "WHERE source_type='ago_mariupol_housing_queue_xlsx' LIMIT 1",
    ).fetchone()
    if not doc:
        return {}
    for row in _rows(doc[1]):
        if not row or not row[0]:
            continue
        total += 1
        by_decree[str(row[2] or "?").strip()] += 1
        if len(row) > 3 and row[3]:
            by_rooms[str(row[3]).strip()] += 1
    log.info("queue %s: %d entries; decree split %s; rooms %s",
             doc[0], total, dict(by_decree), dict(sorted(by_rooms.items())))
    return {"title": doc[0], "total": total,
            "by_decree": dict(by_decree), "by_rooms": dict(by_rooms)}


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    con = sqlite3.connect(config.STATE_DB)
    cur = con.cursor()

    dist = parse_distribution(cur)
    parse_queue(cur)

    out = ROOT / "data" / "parsed" / "ago_lost_dwellings.jsonl"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as f:
        for rec in dist:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # building-level summary for review
    by_bldg = Counter(r["building_id"] for r in dist)
    latest = max((r["list_date"] for r in dist if r["list_date"]), default=None)
    latest_ids = {r["anon_id"] for r in dist if r["list_date"] == latest}
    print(f"\nwrote {len(dist)} lost-dwelling rows across {len(by_bldg)} distinct "
          f"buildings -> {out}")
    print(f"latest snapshot {latest}: {len(latest_ids)} distinct recipients")
    print("top lost-dwelling buildings (distinct apartments recorded lost):")
    for bid, n in by_bldg.most_common(25):
        print(f"  {n:3}  {bid}")


if __name__ == "__main__":
    main()
