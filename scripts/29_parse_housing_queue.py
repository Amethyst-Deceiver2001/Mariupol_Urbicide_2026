#!/usr/bin/env python3
"""Stage 2h: parse the occupation housing queue + distribution XLSX.

Sources (captured by scripts/16_crawl_housing_queue.py):
  - housing_queue_list      (Ochered_*.xlsx)       5,822 rows
        cols: ID | № в очереди | Постановление | Комнатность      (NO address)
  - housing_distribution_list (Raspredelenie_*.xlsx) 1,889 rows
        cols: ID | Адрес утраченного жилья | Район утраченного жилья

The distribution file is the occupation's OWN acknowledgement that named (here
pseudonymised) households lost specific dwellings -- direct claimant-side
evidence for RD4U category A3.6 (loss of access to property), and the join that
confirms ТСЖ «Троянда-М» (пр-т Ленина 86).

PRIVACY (CLAUDE.md hard rule): a (hex ID + lost apartment) pair is sensitive
personal data about a living displaced person. This parser keeps the
BUILDING-level fields (street/house/building_key/district -- a public seized
building, already on our property spine) at top level, and isolates the
person-level fields (hex_id, apartment) under a nested `claimant` object marked
sensitive, so any shared export / DB load can drop `claimant` wholesale and keep
only the non-PII building aggregate. Output lands in data/parsed/ (gitignored);
the per-person detail is NEVER committed and NEVER loaded to a public view.
The building-level loader (db/load.load_housing_distribution) writes only counts
per building -- no hex IDs, no apartment numbers -- to the corroboration table.

Run locally, no network: python3 scripts/29_parse_housing_queue.py
"""
from __future__ import annotations

import json
import logging
import re
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from mariupol_seizures import config, forensics  # noqa: E402
from mariupol_seizures.normalize.address import (  # noqa: E402
    address_to_building_key,
    norm_commas,
)

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed — run: .venv/bin/pip install openpyxl")

log = logging.getLogger(__name__)

PARSED_DIR = config.PROJECT_ROOT / "data" / "parsed"

# Distribution-file district codes -> the district_key convention used across
# the pipeline (ownerless_registry.jsonl, damage_assessment.jsonl).
_DISTRICT_CODE = {
    "ЖР": "zhovtnevy_mariupol",
    "ИР": "ilyichevsky_mariupol",
    "ОР": "ordzhonikidzevsky_mariupol",
    "ПР": "primorsky_mariupol",
}

# "РФ, ДНР, Г. МАРИУПОЛЬ, <street>, Д. <house>, КВ. <apt>" -- strip the fixed
# federal/city prefix, then pull the Д./КВ. tokens off the tail.
_PREFIX_RE = re.compile(r"^\s*РФ\s*,\s*ДНР\s*,\s*Г\.?\s*МАРИУПОЛЬ\s*,\s*", re.I)
_HOUSE_RE = re.compile(r",?\s*Д\.?\s*([0-9]+(?:\s*/\s*[0-9]+)?[А-Яа-яA-Za-z]?)", re.I)
_APT_RE = re.compile(r",?\s*КВ\.?\s*([0-9]+[А-Яа-яA-Za-z]?)", re.I)


def _latest_source(con, source_type: str):
    """Most-recently-captured XLSX of a given source_type (the lists are living
    documents -- take the newest version by capture time)."""
    row = con.execute(
        """SELECT sha256, raw_path, title FROM source_document
           WHERE source_type = ? AND content_type LIKE '%spreadsheet%'
           ORDER BY captured_at DESC LIMIT 1""",
        (source_type,),
    ).fetchone()
    return row  # (sha256, raw_path, title) or None


def _parse_lost_address(raw: str) -> dict:
    """Split a lost-address string into building-level (public) + apartment
    (sensitive) parts and compute the building_key."""
    s = norm_commas(raw or "").strip()
    s = _PREFIX_RE.sub("", s)
    apt = None
    m_apt = _APT_RE.search(s)
    if m_apt:
        apt = m_apt.group(1)
        s = s[: m_apt.start()] + s[m_apt.end():]
    house = None
    m_house = _HOUSE_RE.search(s)
    if m_house:
        house = m_house.group(1).replace(" ", "")
        s = s[: m_house.start()] + s[m_house.end():]
    street = s.strip(" ,").strip() or None
    building_key = address_to_building_key(street, house)
    return {
        "street_raw": street,
        "house_raw": house,
        "apt_raw": apt,
        "building_key": building_key,
    }


def parse_queue(path: Path, sha: str) -> list[dict]:
    ws = openpyxl.load_workbook(path, read_only=True).active
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # header
    out = []
    for r in rows_iter:
        if not r or r[0] is None:
            continue
        out.append({
            "source_sha256": sha,
            "list": "queue",
            # Building-level: none in this file. Person-level: nested+sensitive.
            "claimant": {
                "hex_id": str(r[0]).strip(),
                "queue_position": (str(r[1]).strip() if len(r) > 1 and r[1] is not None else None),
                "decree": (str(r[2]).strip() if len(r) > 2 and r[2] is not None else None),
                "rooms": (str(r[3]).strip() if len(r) > 3 and r[3] is not None else None),
                "_sensitive": True,
            },
        })
    return out


def parse_distribution(path: Path, sha: str) -> list[dict]:
    ws = openpyxl.load_workbook(path, read_only=True).active
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # header
    out = []
    for r in rows_iter:
        if not r or r[0] is None:
            continue
        addr_raw = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        code = str(r[2]).strip() if len(r) > 2 and r[2] is not None else None
        parsed = _parse_lost_address(addr_raw)
        out.append({
            "source_sha256": sha,
            "list": "distribution",
            # ── building-level (public; a seized building on our spine) ──
            "street_raw": parsed["street_raw"],
            "house_raw": parsed["house_raw"],
            "building_key": parsed["building_key"],
            "district_code": code,
            "district_key": _DISTRICT_CODE.get(code),
            # ── person-level (sensitive: hex id + exact apartment) ──
            "claimant": {
                "hex_id": str(r[0]).strip(),
                "apt_raw": parsed["apt_raw"],
                "lost_address_raw": addr_raw,
                "_sensitive": True,
            },
        })
    return out


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-7s %(message)s")
    con = forensics.open_state()
    PARSED_DIR.mkdir(parents=True, exist_ok=True)

    q = _latest_source(con, "housing_queue_list")
    dist = _latest_source(con, "housing_distribution_list")
    if not q and not dist:
        log.error("No housing_queue_list / housing_distribution_list XLSX captured "
                  "-- run scripts/16_crawl_housing_queue.py from the VPS first.")
        sys.exit(1)

    if q:
        sha, raw_path, title = q
        rows = parse_queue(Path(raw_path), sha)
        out = PARSED_DIR / "housing_queue.jsonl"
        with out.open("w", encoding="utf-8") as fh:
            for row in rows:
                fh.write(json.dumps(row, ensure_ascii=False) + "\n")
        log.info("queue: %d rows -> %s (%s)", len(rows), out.name, title)

    if dist:
        sha, raw_path, title = dist
        rows = parse_distribution(Path(raw_path), sha)
        out = PARSED_DIR / "housing_distribution.jsonl"
        with out.open("w", encoding="utf-8") as fh:
            for row in rows:
                fh.write(json.dumps(row, ensure_ascii=False) + "\n")
        keyed = sum(1 for r in rows if r["building_key"])
        per_district = Counter(r["district_key"] for r in rows)
        log.info("distribution: %d rows -> %s (%s)", len(rows), out.name, title)
        log.info("  building_key resolved: %d/%d", keyed, len(rows))
        log.info("  by district: %s", dict(per_district))
        print(f"distribution: {len(rows)} rows, {keyed} with building_key")


if __name__ == "__main__":
    main()
