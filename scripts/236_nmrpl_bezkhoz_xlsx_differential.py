#!/usr/bin/env python3
"""Batch-parse the dated bezkhoz-list XLSX snapshots pulled from @nmrpl (and
the one duplicate on @mariupol_nash) and diff their addresses against what's
currently loaded in the `property`/`seizure_event(stage='registry_inclusion')`
spine.

This is the iterative-decree differential method discussed 2026-07-02: each
of these files is a dated "actual as of <date>" snapshot of the bezkhoz
nonresidential-property list. An address present in an EARLY snapshot but
absent from later ones (and not in the final DB tally) is a candidate for
either (a) court-confirmed transfer/removal (check court_petition +
`decree_kind='removal'` decrees) or (b) owner reclaim before designation
became final. An address that's NEW in a later snapshot but never made the
final DB tally is a candidate for the same disambiguation in the other
direction.

Two distinct list families are mixed in the source channel and handled
separately here (different columns, different populations):
  - "ИЖС" (private house) lists -- simple №/Район/Адрес or №/Адрес tables.
  - "нежилой фонd" (commercial/nonresidential) lists -- №/Объект/Адрес/
    площадь tables, the ones with a clear "актуален на <date>" title cell.

Local-only: reads already-captured files from data/raw/ via source_document,
queries the local Postgres spine read-only. No network, no writes to the DB.
Safe to run directly (not a geoblocked crawl).

Output: data/parsed/nmrpl_bezkhoz_xlsx_differential.jsonl (one row per
address-snapshot pair) + a compact console summary. Read the JSONL only if
you need row-level detail -- the summary is designed to answer "worth a closer
look?" without dumping every row into context.

Run:
    PYTHONPATH=src .venv312/bin/python scripts/236_nmrpl_bezkhoz_xlsx_differential.py
"""
from __future__ import annotations

import json
import logging
import os
import re
import sqlite3
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

import openpyxl  # noqa: E402
import psycopg2  # noqa: E402

from mariupol_seizures.normalize.address import address_to_building_key  # noqa: E402

log = logging.getLogger(__name__)

STATE_DB = ROOT / "data" / "state.sqlite"
OUT = ROOT / "data" / "parsed" / "nmrpl_bezkhoz_xlsx_differential.jsonl"

# (label, snapshot_date ISO, source_document.url) -- snapshot_date taken from
# the filename/title where explicit, else the message date is used as a proxy
# (logged, not guessed silently).
TARGETS = [
    ("izhs_2023-03-27", "2023-03-27", "https://t.me/nmrpl/6177"),
    ("izhs_2023-07-26", "2023-07-26", "https://t.me/nmrpl/10347"),
    ("izhs_dop_2023-XX", None, "https://t.me/nmrpl/12625"),
    ("izhs_dop_2023-09-15", "2023-09-15", "https://t.me/nmrpl/13253"),
    ("nezhiloy_dop_undated", None, "https://t.me/nmrpl/8555"),
    ("nezhiloy_2024-05-16_nmrpl", "2024-05-16", "https://t.me/nmrpl/23279"),
    ("nezhiloy_2024-05-16_nash", "2024-05-16", "https://t.me/mariupol_nash/80072"),
    ("nezhiloy_2024-07-16", "2024-07-16", "https://t.me/nmrpl/25812"),
    ("nezhiloy_2024-07-17", "2024-07-17", "https://t.me/nmrpl/25813"),
    ("nezhiloy_2024-07-23", "2024-07-23", "https://t.me/nmrpl/25971"),
]

_ADDR_HEADER_RE = re.compile(r"адрес", re.I)
_HOUSE_HEADER_RE = re.compile(r"№\s*дома|дом", re.I)


def _raw_path(con, url: str) -> Path | None:
    row = con.execute(
        "SELECT raw_path FROM source_document WHERE url=? AND source_type='telegram_document_media'",
        (url,),
    ).fetchone()
    return Path(row[0]) if row else None


_ADDR_LOOKS_LIKE_RE = re.compile(r"[а-яёА-ЯЁ].{2,40},\s*(?:д\.?\s*)?\d+[а-яёА-ЯЁ]?", re.I)


def _find_header_row(ws) -> tuple[int, dict[str, int]] | None:
    """Scan the first ~10 rows for a header containing an 'адрес' cell.
    Returns (row_index, {col_name: col_index}) or None. Falls back to
    picking the column whose cells most often *look like* an address
    (comma + house number) when no explicit header row exists -- some
    @nmrpl snapshots ship headerless data straight from row 1."""
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows[:10]):
        cols: dict[str, int] = {}
        for j, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            if _ADDR_HEADER_RE.search(cell):
                cols["addr"] = j
            elif _HOUSE_HEADER_RE.search(cell):
                cols["house"] = j
        if "addr" in cols:
            return i, cols

    # fallback: no header -- score columns by address-likeness across the
    # first 30 data rows, pick the best; header_row=-1 means "start at row 0"
    hits: dict[int, int] = {}
    for row in rows[:30]:
        for j, cell in enumerate(row):
            if isinstance(cell, str) and _ADDR_LOOKS_LIKE_RE.search(cell):
                hits[j] = hits.get(j, 0) + 1
    if hits:
        best = max(hits, key=hits.get)
        if hits[best] >= 3:
            return -1, {"addr": best}
    return None


def _extract_addresses(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = _find_header_row(ws)
    if header is None:
        log.warning("no address header found in %s -- skipping", path.name)
        return []
    header_row_idx, cols = header
    addrs: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row_idx:
            continue
        addr_cell = row[cols["addr"]] if cols["addr"] < len(row) else None
        if not addr_cell or not str(addr_cell).strip():
            continue
        addr = str(addr_cell).strip()
        if "house" in cols and cols["house"] < len(row) and row[cols["house"]]:
            addr = f"{addr}, {row[cols['house']]}"
        addrs.append(addr)
    return addrs


_APT_SUFFIX_RE = re.compile(r"\s*,?\s*кв\.?\s*\S+\s*$", re.I)


def _split_street_house(addr: str) -> tuple[str | None, str | None, str | None]:
    """Return (street, house, apt_raw). apt_raw is the stripped 'кв. N'
    suffix if present, so the building key is computed on the building
    address alone -- apartment-level rows must not spawn spurious distinct
    building_ids."""
    apt_raw = None
    m_apt = _APT_SUFFIX_RE.search(addr)
    if m_apt:
        apt_raw = m_apt.group(0).strip(" ,")
        addr = addr[:m_apt.start()].strip()

    parts = [p.strip() for p in addr.split(",")]
    if len(parts) >= 2:
        return parts[0], parts[-1], apt_raw
    m = re.match(r"^(.*?)[,\s]+(?:д\.?\s*)?(\d+[а-яёА-ЯЁ]?(?:/\d+)?)\s*$", addr)
    if m:
        return m.group(1).strip(), m.group(2).strip(), apt_raw
    return addr, None, apt_raw


def _spine_building_keys(pg) -> set[str]:
    cur = pg.cursor()
    cur.execute("SELECT DISTINCT building_id FROM property WHERE building_id IS NOT NULL")
    return {r[0] for r in cur.fetchall()}


def main() -> None:
    con = sqlite3.connect(STATE_DB)
    dsn = os.environ.get("DATABASE_URL", "postgresql://localhost:5433/mariupol_seizures")
    pg = psycopg2.connect(dsn)
    spine_keys = _spine_building_keys(pg)
    log.info("spine has %d distinct building_id values", len(spine_keys))

    fh = OUT.open("w", encoding="utf-8")
    per_file: dict[str, set[str]] = {}
    n_rows_total = 0
    n_unparsed_total = 0

    for label, snap_date, url in TARGETS:
        path = _raw_path(con, url)
        if path is None or not path.exists():
            log.warning("%s: no captured file for %s -- skipping", label, url)
            continue
        addrs = _extract_addresses(path)
        keys: set[str] = set()
        n_unparsed = 0
        for raw in addrs:
            street, house, apt_raw = _split_street_house(raw)
            key = address_to_building_key(street, house) if street else None
            if key is None:
                n_unparsed += 1
                continue
            keys.add(key)
            fh.write(json.dumps({
                "label": label, "snapshot_date": snap_date, "url": url,
                "raw_address": raw, "building_id": key, "apt_raw": apt_raw,
                "on_spine": key in spine_keys,
            }, ensure_ascii=False) + "\n")
        per_file[label] = keys
        n_rows_total += len(addrs)
        n_unparsed_total += n_unparsed
        log.info("%-28s rows=%4d  parsed_keys=%4d  unparsed=%3d  on_spine=%4d",
                  label, len(addrs), len(keys), n_unparsed, len(keys & spine_keys))

    fh.close()

    print(f"\n{'='*76}")
    print("SUMMARY -- addresses per snapshot vs. current spine")
    print(f"{'='*76}")
    for label, keys in per_file.items():
        missing = keys - spine_keys
        print(f"  {label:28s} n={len(keys):4d}  NOT on spine={len(missing):4d}")

    # cross-snapshot differential within the nezhiloy (nonresidential) family,
    # which has a clean date ladder: 16.05.2024 -> 16.07.2024 -> 17.07.2024 -> 23.07.2024
    ladder = [k for k in ["nezhiloy_2024-05-16_nmrpl", "nezhiloy_2024-07-16",
                           "nezhiloy_2024-07-17", "nezhiloy_2024-07-23"] if k in per_file]
    print(f"\n{'='*76}")
    print("NEZHILOY FOND date-ladder differential")
    print(f"{'='*76}")
    for a, b in zip(ladder, ladder[1:]):
        dropped = per_file[a] - per_file[b]
        added = per_file[b] - per_file[a]
        print(f"  {a} -> {b}:  dropped={len(dropped):3d}  added={len(added):3d}")
        for k in sorted(dropped):
            print(f"      DROPPED  {k}  on_spine={k in spine_keys}")
        for k in sorted(added):
            print(f"      ADDED    {k}  on_spine={k in spine_keys}")

    print(f"\n  total rows scanned={n_rows_total}  unparsed_address={n_unparsed_total}")
    print(f"  -> {OUT}")

    pg.close()
    con.close()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    main()
