#!/usr/bin/env python3
"""Stage 2c: forensically capture and parse the Russian damage-assessment XLSX.

Input:  /Users/ak/Downloads/Russian_damage_assessment.xlsx
        (Russian federal reconstruction-tracking document, 1,941 buildings)
Output: data/parsed/damage_assessment.jsonl  — one JSON object per building row

What this file is
-----------------
A Russian occupation tracking spreadsheet assigning named Russian contractors
(Крост, ГК ЕКС, Трансстройинвест, Московский политех, Крокус Групп,
Тульская/Московская oblasti) to damaged buildings in Mariupol.  Fields include:
  - Full address + district (joinable to court_case and seizure_event by address)
  - % разрушения (destruction %) — damage quantum for RD4U claim categories
  - Группа 1-4 (4 = demolish) — group 4 strongly predicts ownerless proceedings
  - Примечание — free-text notes recording direct conflict damage:
    "прямое попадание", "сгорел", "обвал", "труп" — establishes causal chain
    from armed conflict → destruction → seizure

Why this matters
----------------
Three independent occupation sources pointing at the same address:
  1. This file  (Russian federal damage assessment)
  2. Municipal ownerless register  (мэрия designation → court petition)
  3. Court docket  (petition → transfer order)
= legal-grade linkage under CLAUDE.md ≥2-source rule.

Forensic note
-------------
The file is captured to data/raw/ (SHA-256 keyed, with .meta.json sidecar)
before parsing — Berkeley Protocol chain of custody for a document obtained
outside the automated crawl pipeline.
"""
from __future__ import annotations

import json
import logging
import sys
from pathlib import Path

log = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from mariupol_seizures import config, forensics  # noqa: E402

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed — run: pip install openpyxl")

# ── source file ──────────────────────────────────────────────────────────────

SOURCE_PATH = Path(
    __import__("os").environ.get(
        "DAMAGE_ASSESSMENT_XLSX",
        "/Users/ak/Downloads/Russian_damage_assessment.xlsx",
    )
)

# Provenance: where this file came from. Override with DAMAGE_ASSESSMENT_SOURCE_URL.
SOURCE_URL = __import__("os").environ.get(
    "DAMAGE_ASSESSMENT_SOURCE_URL",
    "file:///Users/ak/Downloads/Russian_damage_assessment.xlsx",
)

# ── column indices (0-based) — confirmed from header inspection ───────────────

C_SEQ        = 2
C_CONTRACTOR = 4
C_EXECUTOR   = 5
C_DISTRICT   = 7
C_MICRODISTRICT = 8
C_STREET_TYPE = 9
C_STREET_NAME = 10
C_BUILDING_NO = 12
C_ADDRESS    = 14
C_PHASE      = 15   # Очередность (I / II)
C_PROP_TYPE  = 16   # жилое / нежилое
C_BLDG_CLASS = 17   # МКЖД / ИЖС / school / hospital / etc.
C_FLOORS     = 24
C_ENTRANCES  = 25
C_APARTMENTS = 26
C_GROUP      = 31   # 1=cosm, 2=current, 3=капремонт, 4=demolish
C_NOTES      = 49   # Примечание
C_CONTRACTOR_NOTES = 50
C_DESTRUCTION_PCT  = 53  # % разрушения

# Header rows: rows 0-7 (0-indexed); data starts at row 8.
DATA_START_ROW = 8

# ── district → court key mapping ─────────────────────────────────────────────
# Russian renaming of Ukrainian district names. Must match crawl/courts.py keys.

DISTRICT_KEY: dict[str, str] = {
    "Октябрьский":       "zhovtnevy_mariupol",        # former Жовтневый
    "Орджоникидзевский": "ordzhonikidzevsky_mariupol",
    "Орджоникидзевский ": "ordzhonikidzevsky_mariupol",  # trailing-space variant
    "Ильичевский":       "ilyichevsky_mariupol",
    "Приморский":        "primorsky_mariupol",
}


# ── parse ─────────────────────────────────────────────────────────────────────

def _str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_rows(xlsx_path: Path, source_sha256: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_rows = list(ws.iter_rows(values_only=True))

    records: list[dict] = []
    skipped = 0
    for i, row in enumerate(raw_rows[DATA_START_ROW:], start=DATA_START_ROW):
        seq = row[C_SEQ] if C_SEQ < len(row) else None
        if seq is None or str(seq).strip() == "":
            skipped += 1
            continue
        try:
            seq_int = int(seq)
        except (ValueError, TypeError):
            skipped += 1
            continue

        district_raw = _str(row[C_DISTRICT] if C_DISTRICT < len(row) else None)
        district_key = DISTRICT_KEY.get(district_raw or "", None)
        if district_key is None and district_raw:
            log.warning("unknown district %r at row %d — district_key will be null", district_raw, i)

        records.append({
            "source_sha256":       source_sha256,
            "seq_no":              seq_int,
            "district_raw":        district_raw,
            "district_key":        district_key,
            "microdistrict":       _str(row[C_MICRODISTRICT] if C_MICRODISTRICT < len(row) else None),
            "street_type":         _str(row[C_STREET_TYPE] if C_STREET_TYPE < len(row) else None),
            "street_name":         _str(row[C_STREET_NAME] if C_STREET_NAME < len(row) else None),
            "building_no":         _str(row[C_BUILDING_NO] if C_BUILDING_NO < len(row) else None),
            "address_raw":         _str(row[C_ADDRESS] if C_ADDRESS < len(row) else None),
            "priority_phase":      _str(row[C_PHASE] if C_PHASE < len(row) else None),
            "property_type":       _str(row[C_PROP_TYPE] if C_PROP_TYPE < len(row) else None),
            "building_class":      _str(row[C_BLDG_CLASS] if C_BLDG_CLASS < len(row) else None),
            "floors":              _int(row[C_FLOORS] if C_FLOORS < len(row) else None),
            "entrances":           _str(row[C_ENTRANCES] if C_ENTRANCES < len(row) else None),
            "apartments":          _int(row[C_APARTMENTS] if C_APARTMENTS < len(row) else None),
            "group":               _str(row[C_GROUP] if C_GROUP < len(row) else None),
            "notes":               _str(row[C_NOTES] if C_NOTES < len(row) else None),
            "contractor_notes":    _str(row[C_CONTRACTOR_NOTES] if C_CONTRACTOR_NOTES < len(row) else None),
            "destruction_pct":     _float(row[C_DESTRUCTION_PCT] if C_DESTRUCTION_PCT < len(row) else None),
            "contractor":          _str(row[C_CONTRACTOR] if C_CONTRACTOR < len(row) else None),
            "responsible_executor": _str(row[C_EXECUTOR] if C_EXECUTOR < len(row) else None),
        })

    log.info("parsed %d records (%d skipped)", len(records), skipped)
    return records


# ── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(message)s")

    if not SOURCE_PATH.exists():
        log.error("source file not found: %s", SOURCE_PATH)
        log.error("set DAMAGE_ASSESSMENT_XLSX env var to override the default path")
        sys.exit(1)

    con = forensics.open_state()

    # ── capture forensically ────────────────────────────────────────────────
    raw_bytes = SOURCE_PATH.read_bytes()
    sha = forensics.capture_source(
        raw_bytes,
        url=SOURCE_URL,
        source_type="damage_assessment_xlsx",
        title="Russian occupation damage-assessment and reconstruction-tracking "
              "spreadsheet for Mariupol (1,941 buildings)",
        description=(
            "Russian federal reconstruction-tracking document assigning named "
            "contractors (Крост, ГК ЕКС, Трансстройинвест, Крокус Групп, "
            "Тульская/Московская oblasti) to damaged Mariupol buildings. "
            "Fields: address, district, % разрушения, group (1-4), notes "
            "with direct conflict-damage evidence. Third independent source "
            "for legal-grade address linkage alongside municipal ownerless "
            "register and court docket."
        ),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        http_status=200,
        con=con,
    )
    log.info("captured: sha256=%s", sha)

    # ── parse ───────────────────────────────────────────────────────────────
    records = parse_rows(SOURCE_PATH, sha)

    # ── write output ────────────────────────────────────────────────────────
    out_dir = config.PROJECT_ROOT / "data" / "parsed"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "damage_assessment.jsonl"

    with out_path.open("w", encoding="utf-8") as fh:
        for rec in records:
            fh.write(json.dumps(rec, ensure_ascii=False, default=str) + "\n")

    log.info("wrote %d records to %s", len(records), out_path)

    # ── summary stats ───────────────────────────────────────────────────────
    from collections import Counter
    districts = Counter(r["district_key"] or "UNKNOWN" for r in records)
    groups = Counter(r["group"] or "?" for r in records)
    pct_pop = sum(1 for r in records if r["destruction_pct"] is not None)
    g4_100 = sum(1 for r in records if r["group"] == "4" and r["destruction_pct"] == 100)

    log.info("districts: %s", dict(districts))
    log.info("groups:    %s", dict(sorted(groups.items())))
    log.info("%% разрушения populated: %d/%d", pct_pop, len(records))
    log.info("group-4 + 100%% destruction: %d (direct ownerless candidates)", g4_100)


if __name__ == "__main__":
    main()
